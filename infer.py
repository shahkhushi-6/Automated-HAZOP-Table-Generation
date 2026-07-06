"""
infer.py — Generate a HAZOP table for a new P&ID
==================================================
Loads a trained checkpoint and runs inference over every applicable
(node, guideword) combination in a new system folder, producing an
Excel HAZOP table.

Usage:
    python infer.py \
        --checkpoint ./checkpoints/best_model.pt \
        --system_dir ./new_system \
        --output_file ./hazop_output.xlsx \
        [--guidewords NO_FLOW MORE_FLOW ...] \
        [--num_beams 4]

The system_dir folder needs:
    - nodes.csv
    - adjacency.csv
    - edges.csv
    
It does NOT need hazop.csv (which contains ground truth answers).
"""

import argparse
import logging
import os
from pathlib import Path

import torch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from transformers import AutoTokenizer
from torch_geometric.data import Data as PyGData
from torch_geometric.data import Batch as PyGBatch

from dataset import (
    GUIDEWORDS, GUIDEWORD2IDX, OUTPUT_KEYS,
    NODE_FEAT_DIM, EDGE_FEAT_DIM,
    read_nodes, read_adjacency, read_edges,
)
from model import HAZOPModel

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s",
                    datefmt="%H:%M:%S")


# ──────────────────────────────────────────────────────────────────────────────
# Default guidewords applied to every node
# ──────────────────────────────────────────────────────────────────────────────

DEFAULT_GUIDEWORDS = [
    "NO_FLOW", "MORE_FLOW", "LESS_FLOW", "REVERSE_FLOW",
    "MORE_PRESSURE", "LESS_PRESSURE",
    "MORE_TEMP", "LESS_TEMP",
    "MORE_LEVEL", "LESS_LEVEL",
]


# ──────────────────────────────────────────────────────────────────────────────
# Excel output builder
# ──────────────────────────────────────────────────────────────────────────────

HEADER_COLS = ["Node", "Guideword", "Deviation",
               "Causes", "Consequences", "Safeguards", "Recommendations"]

COL_WIDTHS = [20, 20, 35, 45, 45, 40, 40]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
BORDER_SIDE = Side(style="thin", color="AAAAAA")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                     top=BORDER_SIDE, bottom=BORDER_SIDE)
WRAP = Alignment(wrap_text=True, vertical="top")


def save_excel(rows: list, output_path: str):
    """
    Create Excel workbook with HAZOP table.

    Args:
        rows: list of lists, one per HAZOP row
        output_path: output file path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HAZOP Table"

    # Header row
    for col_idx, (header, width) in enumerate(
            zip(HEADER_COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = CELL_BORDER
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = WRAP
            cell.border = CELL_BORDER

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(output_path)
    log.info("Saved HAZOP table → %s  (%d rows)", output_path, len(rows))


# ──────────────────────────────────────────────────────────────────────────────
# Graph builder (for inference systems without hazop.csv)
# ──────────────────────────────────────────────────────────────────────────────

def build_inference_graph(system_dir: str) -> tuple:
    """
    Build graph for inference system (no hazop.csv required).

    Args:
        system_dir: path to system folder

    Returns:
        (graph, node_ids)

    Raises:
        ValueError if any required file is missing or invalid
    """
    sys_dir = Path(system_dir)

    nodes_path = sys_dir / "nodes.csv"
    adj_path = sys_dir / "adjacency.csv"
    edges_path = sys_dir / "edges.csv"

    for p in [nodes_path, adj_path, edges_path]:
        if not p.exists():
            raise ValueError(f"Missing required file: {p}")

    try:
        # Read nodes
        node_ids, node_feat = read_nodes(str(nodes_path))
        node2idx = {nid: i for i, nid in enumerate(node_ids)}

        # Read adjacency
        edge_pairs = read_adjacency(str(adj_path), node2idx)

        # Read edge features
        edge_feats = read_edges(str(edges_path))

    except Exception as e:
        raise ValueError(f"Error reading graph files: {e}")

    # If no edges, create self-loops
    if not edge_pairs:
        n = node_feat.size(0)
        edge_pairs = [(i, i) for i in range(n)]
        edge_feats = [[0.0] * EDGE_FEAT_DIM] * n
        log.warning("No edges found; using self-loops for %d nodes", n)

    # Align edges and features
    n_edges = min(len(edge_pairs), len(edge_feats))
    edge_pairs = edge_pairs[:n_edges]
    edge_feats = edge_feats[:n_edges]

    src_nodes = [e[0] for e in edge_pairs]
    dst_nodes = [e[1] for e in edge_pairs]

    graph = PyGData(
        x=node_feat,
        edge_index=torch.tensor([src_nodes, dst_nodes], dtype=torch.long),
        edge_attr=torch.tensor(edge_feats, dtype=torch.float),
    )

    log.info("Built graph: %d nodes, %d edges", len(node_ids), n_edges)
    return graph, node_ids


# ──────────────────────────────────────────────────────────────────────────────
# Main inference routine
# ──────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="HAZOP inference")
    parser.add_argument("--checkpoint", required=True,
                        help="Path to best_model.pt")
    parser.add_argument("--system_dir", required=True,
                        help="Folder with nodes.csv / adjacency.csv / edges.csv")
    parser.add_argument("--output_file", default="hazop_output.xlsx")
    parser.add_argument("--guidewords", nargs="+",
                        default=DEFAULT_GUIDEWORDS,
                        help="Guidewords to apply to every node")
    parser.add_argument("--num_beams", type=int, default=4)
    parser.add_argument("--max_output_len", type=int, default=128)
    args = parser.parse_args()

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    log.info("Device: %s", device)

    # ── Load checkpoint ──────────────────────────────────────────────────
    log.info("Loading checkpoint: %s", args.checkpoint)
    if not os.path.exists(args.checkpoint):
        raise ValueError(f"Checkpoint not found: {args.checkpoint}")

    ckpt = torch.load(args.checkpoint, map_location=device)
    train_args = ckpt["args"]

    log.info("Checkpoint from epoch %d (val_loss=%.4f)",
             ckpt["epoch"], ckpt["val_loss"])

    # Load tokenizer
    tokenizer = AutoTokenizer.from_pretrained(train_args["t5_model"])

    # Rebuild model with same hyperparameters
    model = HAZOPModel(
        node_feat_dim=NODE_FEAT_DIM,
        gnn_hidden=train_args["gnn_hidden"],
        gnn_out=train_args["gnn_out"],
        gnn_layers=train_args["gnn_layers"],
        num_guidewords=len(GUIDEWORDS),
        gw_embed_dim=train_args["gw_embed_dim"],
        t5_model_name=train_args["t5_model"],
        output_keys=OUTPUT_KEYS,
    ).to(device)

    model.load_state_dict(ckpt["model_state"])
    model.eval()
    log.info("Model loaded and ready for inference")

    # ── Build graph ──────────────────────────────────────────────────────
    log.info("Building graph from: %s", args.system_dir)
    try:
        graph, node_ids = build_inference_graph(args.system_dir)
    except ValueError as e:
        log.error("Failed to build graph: %s", e)
        return

    # ── Validate requested guidewords ────────────────────────────────────
    valid_gw = []
    for gw in args.guidewords:
        gw_upper = gw.upper()
        if gw_upper in GUIDEWORD2IDX:
            valid_gw.append(gw_upper)
        else:
            log.warning("Unknown guideword '%s' — skipped", gw)

    if not valid_gw:
        log.error("No valid guidewords specified. Available: %s",
                 list(GUIDEWORD2IDX.keys()))
        return

    # ── Inference loop ───────────────────────────────────────────────────
    rows = []
    total = len(node_ids) * len(valid_gw)
    done = 0

    log.info("Generating %d HAZOP rows (%d nodes × %d guidewords) …",
             total, len(node_ids), len(valid_gw))

    for node_idx, node_id in enumerate(node_ids):
        for gw in valid_gw:
            gw_idx = GUIDEWORD2IDX[gw]

            result = model.generate_hazop_row(
                graph=graph,
                node_idx=node_idx,
                guideword_idx=gw_idx,
                tokenizer=tokenizer,
                max_output_len=args.max_output_len,
                num_beams=args.num_beams,
            )

            row = [
                node_id,
                gw,
                result.get("deviation", ""),
                result.get("causes", ""),
                result.get("consequences", ""),
                result.get("safeguards", ""),
                result.get("recommendations", ""),
            ]
            rows.append(row)
            done += 1

            if done % 10 == 0 or done == total:
                log.info("  %d / %d rows complete", done, total)

    # ── Save Excel output ────────────────────────────────────────────────
    os.makedirs(os.path.dirname(args.output_file) or ".", exist_ok=True)
    save_excel(rows, args.output_file)
    log.info("HAZOP generation complete!")


if __name__ == "__main__":
    main()
